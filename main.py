import matplotlib
matplotlib.use('TkAgg')

import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as ttkb
from ttkbootstrap.constants import *
import database
import db_operations
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import difflib

class WeddingScriptApp:
    def __init__(self, root):
        self.root = root
        self.root.title("婚礼主持词版本校对记录器")
        self.root.geometry("1400x900")
        self.root.minsize(1200, 700)
        
        database.init_database()
        
        self.current_script_id = None
        self.setup_ui()
        self.load_script_list()
    
    def setup_ui(self):
        style = ttkb.Style(theme='cosmo')
        
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        left_frame = ttk.Frame(main_frame, width=350)
        left_frame.pack(side=LEFT, fill=BOTH, padx=(0, 5))
        
        right_frame = ttk.Frame(main_frame)
        right_frame.pack(side=RIGHT, fill=BOTH, expand=True, padx=(5, 0))
        
        self.setup_left_panel(left_frame)
        self.setup_right_panel(right_frame)
    
    def setup_left_panel(self, parent):
        header_frame = ttk.Frame(parent)
        header_frame.pack(fill=X, pady=(0, 10))
        
        title_label = ttk.Label(header_frame, text="主持词列表", font=('Microsoft YaHei', 16, 'bold'))
        title_label.pack(side=LEFT)
        
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(header_frame, textvariable=self.search_var, width=20)
        search_entry.pack(side=RIGHT, padx=(10, 0))
        
        search_btn = ttk.Button(header_frame, text="搜索", command=self.search_scripts)
        search_btn.pack(side=RIGHT)
        
        filter_frame = ttk.LabelFrame(parent, text="筛选条件")
        filter_frame.pack(fill=X, pady=(0, 10))
        
        ttk.Label(filter_frame, text="主持人:").grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.host_combobox = ttk.Combobox(filter_frame, width=15)
        self.host_combobox.grid(row=0, column=1, padx=5, pady=5)
        self.load_host_names()
        
        ttk.Label(filter_frame, text="定稿状态:").grid(row=0, column=2, padx=5, pady=5, sticky=W)
        self.status_combobox = ttk.Combobox(filter_frame, width=10, values=['', '未定稿', '已定稿'])
        self.status_combobox.grid(row=0, column=3, padx=5, pady=5)
        
        filter_btn = ttk.Button(filter_frame, text="筛选", command=self.filter_scripts)
        filter_btn.grid(row=0, column=4, padx=5, pady=5)
        
        clear_btn = ttk.Button(filter_frame, text="清空", command=self.clear_filter)
        clear_btn.grid(row=0, column=5, padx=5, pady=5)
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=BOTH, expand=True)
        
        columns = ('id', 'record_no', 'bride', 'groom', 'date', 'host', 'version', 'round', 'status')
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=20)
        self.tree.heading('record_no', text='记录编号')
        self.tree.heading('bride', text='新娘姓名')
        self.tree.heading('groom', text='新郎姓名')
        self.tree.heading('date', text='婚礼日期')
        self.tree.heading('host', text='主持人')
        self.tree.heading('version', text='版本号')
        self.tree.heading('round', text='反馈轮次')
        self.tree.heading('status', text='定稿状态')
        
        self.tree.column('id', width=0, stretch=NO)
        self.tree.column('record_no', width=100)
        self.tree.column('bride', width=70)
        self.tree.column('groom', width=70)
        self.tree.column('date', width=80)
        self.tree.column('host', width=70)
        self.tree.column('version', width=60)
        self.tree.column('round', width=60)
        self.tree.column('status', width=60)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.tree.pack(fill=BOTH, expand=True)
        
        self.tree.bind('<<TreeviewSelect>>', self.on_script_select)
        
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=X, pady=(10, 0))
        
        self.add_btn = ttk.Button(btn_frame, text="新增记录", command=self.open_add_dialog)
        self.add_btn.pack(side=LEFT, padx=(0, 5), fill=X, expand=True)
        
        self.edit_btn = ttk.Button(btn_frame, text="编辑记录", command=self.open_edit_dialog, state=DISABLED)
        self.edit_btn.pack(side=LEFT, padx=5, fill=X, expand=True)
        
        self.delete_btn = ttk.Button(btn_frame, text="删除记录", command=self.delete_script, state=DISABLED)
        self.delete_btn.pack(side=LEFT, padx=5, fill=X, expand=True)
        
        self.export_btn = ttk.Button(btn_frame, text="导出报表", command=self.export_report)
        self.export_btn.pack(side=LEFT, padx=(5, 0), fill=X, expand=True)
    
    def setup_right_panel(self, parent):
        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=BOTH, expand=True)
        
        self.detail_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.detail_frame, text="详情")
        
        self.change_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.change_frame, text="变更记录")
        
        self.version_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.version_frame, text="版本管理")
        
        self.approval_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.approval_frame, text="审批流程")
        
        self.stats_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.stats_frame, text="统计分析")
        
        self.setup_detail_panel(self.detail_frame)
        self.setup_change_panel(self.change_frame)
        self.setup_version_panel(self.version_frame)
        self.setup_approval_panel(self.approval_frame)
        self.setup_stats_panel(self.stats_frame)
    
    def setup_detail_panel(self, parent):
        detail_frame = ttk.LabelFrame(parent, text="主持词详情")
        detail_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        fields = [
            ('记录编号', 'record_no'),
            ('新娘姓名', 'bride_name'),
            ('新郎姓名', 'groom_name'),
            ('婚礼日期', 'wedding_date'),
            ('主持人', 'host_name'),
            ('当前版本', 'current_version'),
            ('反馈轮次', 'feedback_round'),
            ('定稿状态', 'finalized_status'),
            ('备注', 'remarks')
        ]
        
        self.detail_vars = {}
        for i, (label, key) in enumerate(fields):
            ttk.Label(detail_frame, text=f"{label}:").grid(row=i, column=0, padx=10, pady=8, sticky=W)
            var = tk.StringVar()
            self.detail_vars[key] = var
            if key == 'remarks':
                text = ttk.Label(detail_frame, textvariable=var, wraplength=500, justify=LEFT)
                text.grid(row=i, column=1, padx=10, pady=8, sticky=W)
            else:
                ttk.Label(detail_frame, textvariable=var).grid(row=i, column=1, padx=10, pady=8, sticky=W)
        
        content_frame = ttk.LabelFrame(parent, text="主持词正文")
        content_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        self.content_text = tk.Text(content_frame, wrap=tk.WORD, font=('Microsoft YaHei', 11))
        self.content_text.pack(fill=BOTH, expand=True, padx=10, pady=10)
        self.content_text.config(state=tk.DISABLED)
        
        no_data_label = ttk.Label(detail_frame, text="请选择一条主持词记录查看详情", font=('Microsoft YaHei', 14))
        no_data_label.grid(row=0, column=0, columnspan=2, padx=20, pady=100)
        self.no_data_label = no_data_label
    
    def setup_change_panel(self, parent):
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=X, padx=10, pady=10)
        
        self.add_change_btn = ttk.Button(btn_frame, text="新增变更记录", command=self.open_add_change_dialog, state=DISABLED)
        self.add_change_btn.pack(side=LEFT)
        
        self.delete_change_btn = ttk.Button(btn_frame, text="删除变更记录", command=self.delete_selected_change, state=DISABLED)
        self.delete_change_btn.pack(side=LEFT, padx=10)
        
        tree_frame = ttk.Frame(parent)
        tree_frame.pack(fill=BOTH, expand=True, padx=10)
        
        columns = ('id', 'modify_date', 'paragraph', 'reason', 'source', 'adopted')
        self.change_tree = ttk.Treeview(tree_frame, columns=columns, show='headings', height=15)
        self.change_tree.heading('modify_date', text='修改日期')
        self.change_tree.heading('paragraph', text='修改段落')
        self.change_tree.heading('reason', text='修改原因')
        self.change_tree.heading('source', text='反馈来源')
        self.change_tree.heading('adopted', text='是否采纳')
        
        self.change_tree.column('id', width=0, stretch=NO)
        self.change_tree.column('modify_date', width=150)
        self.change_tree.column('paragraph', width=120)
        self.change_tree.column('reason', width=150)
        self.change_tree.column('source', width=100)
        self.change_tree.column('adopted', width=80)
        
        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=self.change_tree.yview)
        self.change_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.change_tree.pack(fill=BOTH, expand=True)
        
        self.change_tree.bind('<<TreeviewSelect>>', self.on_change_select)
    
    def setup_version_panel(self, parent):
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=X, padx=10, pady=10)
        
        self.compare_btn = ttk.Button(btn_frame, text="差异对比", command=self.open_compare_dialog, state=DISABLED)
        self.compare_btn.pack(side=LEFT)
        
        self.rollback_btn = ttk.Button(btn_frame, text="版本回滚", command=self.open_rollback_dialog, state=DISABLED)
        self.rollback_btn.pack(side=LEFT, padx=10)
        
        self.branch_btn = ttk.Button(btn_frame, text="复制为新版本分支", command=self.open_branch_dialog, state=DISABLED)
        self.branch_btn.pack(side=LEFT, padx=10)
        
        self.export_diff_btn = ttk.Button(btn_frame, text="导出差异摘要", command=self.export_diff_summary, state=DISABLED)
        self.export_diff_btn.pack(side=LEFT, padx=10)
        
        version_frame = ttk.LabelFrame(parent, text="版本历史")
        version_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        columns = ('id', 'version', 'label', 'status', 'created_at')
        self.version_tree = ttk.Treeview(version_frame, columns=columns, show='headings', height=12)
        self.version_tree.heading('version', text='版本号')
        self.version_tree.heading('label', text='版本标签')
        self.version_tree.heading('status', text='状态')
        self.version_tree.heading('created_at', text='创建时间')
        
        self.version_tree.column('id', width=0, stretch=NO)
        self.version_tree.column('version', width=80)
        self.version_tree.column('label', width=150)
        self.version_tree.column('status', width=80)
        self.version_tree.column('created_at', width=160)
        
        scrollbar = ttk.Scrollbar(version_frame, orient=VERTICAL, command=self.version_tree.yview)
        self.version_tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.version_tree.pack(fill=BOTH, expand=True)
        
        self.version_tree.bind('<<TreeviewSelect>>', self.on_version_select)
        
        preview_frame = ttk.LabelFrame(parent, text="版本内容预览")
        preview_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        self.version_preview = tk.Text(preview_frame, wrap=tk.WORD, font=('Microsoft YaHei', 10))
        self.version_preview.pack(fill=BOTH, expand=True, padx=10, pady=10)
        self.version_preview.config(state=tk.DISABLED)
        
        rollback_frame = ttk.LabelFrame(parent, text="回滚记录")
        rollback_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        rollback_columns = ('id', 'from_version', 'to_version', 'reason', 'operated_by', 'operated_at')
        self.rollback_tree = ttk.Treeview(rollback_frame, columns=rollback_columns, show='headings', height=6)
        self.rollback_tree.heading('from_version', text='原版本')
        self.rollback_tree.heading('to_version', text='目标版本')
        self.rollback_tree.heading('reason', text='回滚原因')
        self.rollback_tree.heading('operated_by', text='操作人')
        self.rollback_tree.heading('operated_at', text='操作时间')
        
        self.rollback_tree.column('id', width=0, stretch=NO)
        self.rollback_tree.column('from_version', width=80)
        self.rollback_tree.column('to_version', width=80)
        self.rollback_tree.column('reason', width=200)
        self.rollback_tree.column('operated_by', width=80)
        self.rollback_tree.column('operated_at', width=160)
        
        rollback_scrollbar = ttk.Scrollbar(rollback_frame, orient=VERTICAL, command=self.rollback_tree.yview)
        self.rollback_tree.configure(yscrollcommand=rollback_scrollbar.set)
        rollback_scrollbar.pack(side=RIGHT, fill=Y)
        self.rollback_tree.pack(fill=BOTH, expand=True)
    
    def setup_approval_panel(self, parent):
        btn_frame = ttk.Frame(parent)
        btn_frame.pack(fill=X, padx=10, pady=10)
        
        self.submit_approval_btn = ttk.Button(btn_frame, text="提交审批", command=self.open_submit_approval_dialog, state=DISABLED)
        self.submit_approval_btn.pack(side=LEFT)
        
        self.host_approve_btn = ttk.Button(btn_frame, text="主持人确认", command=self.open_host_approve_dialog, state=DISABLED)
        self.host_approve_btn.pack(side=LEFT, padx=10)
        
        self.planner_approve_btn = ttk.Button(btn_frame, text="策划师确认", command=self.open_planner_approve_dialog, state=DISABLED)
        self.planner_approve_btn.pack(side=LEFT, padx=10)
        
        self.customer_confirm_btn = ttk.Button(btn_frame, text="客户确认", command=self.open_customer_confirm_dialog, state=DISABLED)
        self.customer_confirm_btn.pack(side=LEFT, padx=10)
        
        self.export_approval_btn = ttk.Button(btn_frame, text="导出审批历史", command=self.export_approval_history, state=DISABLED)
        self.export_approval_btn.pack(side=RIGHT)
        
        status_frame = ttk.LabelFrame(parent, text="审批状态")
        status_frame.pack(fill=X, padx=10, pady=5)
        
        self.approval_status_vars = {}
        status_fields = [
            ('当前状态', 'status'),
            ('当前审批人', 'current_approver'),
            ('客户查看时间', 'customer_view_time'),
            ('确认结果', 'customer_confirm_result'),
            ('退回意见', 'customer_feedback'),
            ('确认人', 'customer_confirmed_by'),
            ('确认时间', 'customer_confirmed_at'),
            ('提交时间', 'created_at')
        ]
        
        for i, (label, key) in enumerate(status_fields):
            ttk.Label(status_frame, text=f"{label}:").grid(row=i, column=0, padx=10, pady=5, sticky=W)
            var = tk.StringVar(value='-')
            self.approval_status_vars[key] = var
            if key == 'customer_feedback':
                text = ttk.Label(status_frame, textvariable=var, wraplength=500, justify=LEFT)
                text.grid(row=i, column=1, padx=10, pady=5, sticky=W)
            else:
                ttk.Label(status_frame, textvariable=var).grid(row=i, column=1, padx=10, pady=5, sticky=W)
        
        history_frame = ttk.LabelFrame(parent, text="审批历史")
        history_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        history_columns = ('id', 'action', 'role', 'operator', 'remark', 'created_at')
        self.approval_history_tree = ttk.Treeview(history_frame, columns=history_columns, show='headings', height=8)
        self.approval_history_tree.heading('action', text='操作')
        self.approval_history_tree.heading('role', text='角色')
        self.approval_history_tree.heading('operator', text='操作人')
        self.approval_history_tree.heading('remark', text='备注')
        self.approval_history_tree.heading('created_at', text='时间')
        
        self.approval_history_tree.column('id', width=0, stretch=NO)
        self.approval_history_tree.column('action', width=120)
        self.approval_history_tree.column('role', width=80)
        self.approval_history_tree.column('operator', width=80)
        self.approval_history_tree.column('remark', width=200)
        self.approval_history_tree.column('created_at', width=160)
        
        history_scrollbar = ttk.Scrollbar(history_frame, orient=VERTICAL, command=self.approval_history_tree.yview)
        self.approval_history_tree.configure(yscrollcommand=history_scrollbar.set)
        history_scrollbar.pack(side=RIGHT, fill=Y)
        self.approval_history_tree.pack(fill=BOTH, expand=True)
        
        task_frame = ttk.LabelFrame(parent, text="反馈任务")
        task_frame.pack(fill=BOTH, expand=True, padx=10, pady=5)
        
        task_columns = ('id', 'paragraph', 'feedback', 'status', 'assigned_to', 'created_at', 'completed_at')
        self.feedback_task_tree = ttk.Treeview(task_frame, columns=task_columns, show='headings', height=6)
        self.feedback_task_tree.heading('paragraph', text='修改段落')
        self.feedback_task_tree.heading('feedback', text='反馈内容')
        self.feedback_task_tree.heading('status', text='状态')
        self.feedback_task_tree.heading('assigned_to', text='指派给')
        self.feedback_task_tree.heading('created_at', text='创建时间')
        self.feedback_task_tree.heading('completed_at', text='完成时间')
        
        self.feedback_task_tree.column('id', width=0, stretch=NO)
        self.feedback_task_tree.column('paragraph', width=100)
        self.feedback_task_tree.column('feedback', width=250)
        self.feedback_task_tree.column('status', width=80)
        self.feedback_task_tree.column('assigned_to', width=80)
        self.feedback_task_tree.column('created_at', width=120)
        self.feedback_task_tree.column('completed_at', width=120)
        
        task_scrollbar = ttk.Scrollbar(task_frame, orient=VERTICAL, command=self.feedback_task_tree.yview)
        self.feedback_task_tree.configure(yscrollcommand=task_scrollbar.set)
        task_scrollbar.pack(side=RIGHT, fill=Y)
        self.feedback_task_tree.pack(fill=BOTH, expand=True)
        
        self.complete_task_btn = ttk.Button(task_frame, text="标记完成", command=self.complete_feedback_task, state=DISABLED)
        self.complete_task_btn.pack(side=LEFT, padx=10, pady=5)
        
        self.feedback_task_tree.bind('<<TreeviewSelect>>', self.on_task_select)
    
    def on_task_select(self, event):
        selection = self.feedback_task_tree.selection()
        if selection:
            item = self.feedback_task_tree.item(selection[0])
            status = item['values'][3]
            self.complete_task_btn.config(state=NORMAL if status == '待处理' else DISABLED)
        else:
            self.complete_task_btn.config(state=DISABLED)
    
    def complete_feedback_task(self):
        selection = self.feedback_task_tree.selection()
        if selection:
            item = self.feedback_task_tree.item(selection[0])
            task_id = item['values'][0]
            
            if messagebox.askyesno("确认完成", "确定要标记此任务为已完成吗？"):
                success, error = db_operations.update_feedback_task_status(task_id, '已完成')
                if success:
                    messagebox.showinfo("成功", "任务已标记为完成")
                    self.load_feedback_tasks(self.current_script_id)
                else:
                    messagebox.showerror("错误", f"操作失败: {error}")
    
    def open_submit_approval_dialog(self):
        if self.current_script_id:
            script = db_operations.get_host_script_by_id(self.current_script_id)
            if script:
                dialog = SubmitApprovalDialog(self.root, self, self.current_script_id, script[6])
                self.root.wait_window(dialog.top)
    
    def open_host_approve_dialog(self):
        if self.current_script_id:
            script = db_operations.get_host_script_by_id(self.current_script_id)
            if script:
                dialog = ApproveDialog(self.root, self, self.current_script_id, script[6], '主持人')
                self.root.wait_window(dialog.top)
    
    def open_planner_approve_dialog(self):
        if self.current_script_id:
            script = db_operations.get_host_script_by_id(self.current_script_id)
            if script:
                dialog = ApproveDialog(self.root, self, self.current_script_id, script[6], '策划师')
                self.root.wait_window(dialog.top)
    
    def open_customer_confirm_dialog(self):
        if self.current_script_id:
            script = db_operations.get_host_script_by_id(self.current_script_id)
            if script:
                dialog = CustomerConfirmDialog(self.root, self, self.current_script_id, script[6])
                self.root.wait_window(dialog.top)
    
    def export_approval_history(self):
        if self.current_script_id:
            filename, error = db_operations.export_approval_history(self.current_script_id)
            if filename:
                messagebox.showinfo("成功", f"审批历史已导出为: {filename}")
            else:
                messagebox.showerror("错误", f"导出失败: {error}")
    
    def load_approval_info(self, script_id):
        script = db_operations.get_host_script_by_id(script_id)
        if not script:
            return
        
        version_number = script[6]
        flow = db_operations.get_approval_flow_by_script_and_version(script_id, version_number)
        
        for key in self.approval_status_vars:
            self.approval_status_vars[key].set('-')
        
        if flow:
            status_map = {
                'status': 3, 'current_approver': 4, 'customer_view_time': 5,
                'customer_confirm_result': 6, 'customer_feedback': 7, 'customer_confirmed_by': 8,
                'customer_confirmed_at': 9, 'created_at': 10
            }
            for key, idx in status_map.items():
                if idx < len(flow) and flow[idx]:
                    self.approval_status_vars[key].set(flow[idx])
        
        self.load_approval_history(script_id)
        self.load_feedback_tasks(script_id)
        
        self.update_approval_buttons(script_id, version_number, flow)
    
    def update_approval_buttons(self, script_id, version_number, flow):
        is_finalized = script_id and db_operations.get_host_script_by_id(script_id)[8] == '已定稿'
        
        if is_finalized:
            self.submit_approval_btn.config(state=DISABLED)
            self.host_approve_btn.config(state=DISABLED)
            self.planner_approve_btn.config(state=DISABLED)
            self.customer_confirm_btn.config(state=DISABLED)
            return
        
        self.submit_approval_btn.config(state=NORMAL)
        self.export_approval_btn.config(state=NORMAL)
        
        if flow:
            status = flow[3]
            current_approver = flow[4]
            
            self.host_approve_btn.config(state=NORMAL if current_approver == '主持人' else DISABLED)
            self.planner_approve_btn.config(state=NORMAL if current_approver == '策划师' else DISABLED)
            self.customer_confirm_btn.config(state=NORMAL if current_approver == '客户' else DISABLED)
        else:
            self.host_approve_btn.config(state=DISABLED)
            self.planner_approve_btn.config(state=DISABLED)
            self.customer_confirm_btn.config(state=DISABLED)
    
    def load_approval_history(self, script_id):
        for item in self.approval_history_tree.get_children():
            self.approval_history_tree.delete(item)
        
        histories = db_operations.get_approval_history_by_script_id(script_id)
        for h in histories:
            self.approval_history_tree.insert('', END, values=(h[0], h[4], h[5], h[6], h[7], h[8]))
    
    def load_feedback_tasks(self, script_id):
        for item in self.feedback_task_tree.get_children():
            self.feedback_task_tree.delete(item)
        
        tasks = db_operations.get_feedback_tasks_by_script_id(script_id)
        for task in tasks:
            self.feedback_task_tree.insert('', END, values=(task[0], task[4], task[5], task[6], 
                                                              task[7], task[8], task[9]))
    
    def setup_stats_panel(self, parent):
        for widget in parent.winfo_children():
            widget.destroy()
        
        self.overview_frame = ttk.LabelFrame(parent, text="概览统计")
        self.overview_frame.pack(fill=X, padx=10, pady=10)
        
        self.stat_labels = {}
        stats = db_operations.get_stats_overview()
        stat_items = [
            ('主持总数', stats['total']),
            ('已定稿', stats['finalized']),
            ('平均反馈轮次', stats['avg_round']),
            ('变更记录总数', stats['total_changes'])
        ]
        
        for i, (label, value) in enumerate(stat_items):
            frame = ttk.Frame(self.overview_frame)
            frame.pack(side=LEFT, padx=20, pady=10)
            ttk.Label(frame, text=label, font=('Microsoft YaHei', 10)).pack()
            val_label = ttk.Label(frame, text=str(value), font=('Microsoft YaHei', 20, 'bold'))
            val_label.pack()
            self.stat_labels[label] = val_label
        
        self.chart_frame = ttk.LabelFrame(parent, text="统计图表")
        self.chart_frame.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        self.update_charts()
        
        self.high_freq_frame = ttk.LabelFrame(parent, text="高频问题（出现≥3次）")
        self.high_freq_frame.pack(fill=X, padx=10, pady=10)
        
        self.update_high_freq_issues()
        
        self.paragraph_frame = ttk.LabelFrame(parent, text="修改段落统计")
        self.paragraph_frame.pack(fill=X, padx=10, pady=10)
        
        self.update_paragraph_stats()
    
    def update_paragraph_stats(self):
        for widget in self.paragraph_frame.winfo_children():
            widget.destroy()
        
        paragraph_stats = db_operations.get_modify_paragraph_stats()
        
        if paragraph_stats:
            columns = ('paragraph', 'count')
            tree = ttk.Treeview(self.paragraph_frame, columns=columns, show='headings', height=6)
            tree.heading('paragraph', text='修改段落')
            tree.heading('count', text='修改次数')
            
            tree.column('paragraph', width=300)
            tree.column('count', width=100)
            
            for p in paragraph_stats:
                tree.insert('', END, values=(p[0], p[1]))
            
            tree.pack(fill=X)
        else:
            ttk.Label(self.paragraph_frame, text="暂无修改段落数据").pack(pady=10)
    
    def update_charts(self):
        for widget in self.chart_frame.winfo_children():
            widget.destroy()
        
        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(10, 4))
        
        version_stats = db_operations.get_version_stats()
        if version_stats:
            versions = [str(v[0]) for v in version_stats]
            counts = [v[1] for v in version_stats]
            sns.barplot(x=versions, y=counts, ax=ax1)
            ax1.set_title('版本分布')
            ax1.set_xlabel('版本号')
            ax1.set_ylabel('数量')
        
        reason_stats = db_operations.get_change_reason_stats()
        if reason_stats:
            reasons = [r[0][:8] if len(r[0]) > 8 else r[0] for r in reason_stats]
            counts = [r[1] for r in reason_stats]
            sns.barplot(x=counts, y=reasons, ax=ax2, orient='h')
            ax2.set_title('修改原因TOP10')
            ax2.set_xlabel('数量')
            ax2.set_ylabel('原因')
        
        plt.tight_layout()
        canvas = FigureCanvasTkAgg(fig, master=self.chart_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=BOTH, expand=True)
    
    def update_high_freq_issues(self):
        for widget in self.high_freq_frame.winfo_children():
            widget.destroy()
        
        high_freq_issues = db_operations.get_high_freq_issues()
        high_freq_issues = [i for i in high_freq_issues if i[1] >= 3]
        
        if high_freq_issues:
            columns = ('issue', 'count', 'first', 'last')
            tree = ttk.Treeview(self.high_freq_frame, columns=columns, show='headings', height=5)
            tree.heading('issue', text='问题描述')
            tree.heading('count', text='出现次数')
            tree.heading('first', text='首次出现')
            tree.heading('last', text='最近出现')
            
            tree.column('issue', width=250)
            tree.column('count', width=80)
            tree.column('first', width=120)
            tree.column('last', width=120)
            
            for issue in high_freq_issues:
                tree.insert('', END, values=(issue[0], issue[1], issue[2], issue[3]))
            
            tree.pack(fill=X)
        else:
            ttk.Label(self.high_freq_frame, text="暂无高频问题").pack(pady=10)
    
    def refresh_stats(self):
        stats = db_operations.get_stats_overview()
        stat_items = [
            ('主持总数', stats['total']),
            ('已定稿', stats['finalized']),
            ('平均反馈轮次', stats['avg_round']),
            ('变更记录总数', stats['total_changes'])
        ]
        
        for label, value in stat_items:
            if label in self.stat_labels:
                self.stat_labels[label].config(text=str(value))
        
        self.update_charts()
        self.update_high_freq_issues()
        self.update_paragraph_stats()
    
    def load_script_list(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        scripts = db_operations.get_all_host_scripts()
        for script in scripts:
            self.tree.insert('', END, values=(script[0], script[1], script[2], script[3], script[4], 
                                              script[5], script[6], script[7], script[8]))
    
    def load_host_names(self):
        hosts = db_operations.get_host_names()
        host_list = [''] + [h[0] for h in hosts]
        self.host_combobox['values'] = host_list
    
    def on_script_select(self, event):
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            script_id = item['values'][0]
            self.current_script_id = script_id
            self.load_script_detail(script_id)
            self.load_change_records(script_id)
            self.load_version_history(script_id)
            self.load_rollback_records(script_id)
            self.load_approval_info(script_id)
            
            script = db_operations.get_host_script_by_id(script_id)
            is_finalized = script and script[8] == '已定稿'
            
            self.edit_btn.config(state=DISABLED if is_finalized else NORMAL)
            self.delete_btn.config(state=NORMAL)
            self.add_change_btn.config(state=DISABLED if is_finalized else NORMAL)
            self.compare_btn.config(state=NORMAL)
            self.rollback_btn.config(state=NORMAL)
            self.branch_btn.config(state=NORMAL if is_finalized else DISABLED)
            self.export_diff_btn.config(state=NORMAL)
        else:
            self.current_script_id = None
            self.edit_btn.config(state=DISABLED)
            self.delete_btn.config(state=DISABLED)
            self.add_change_btn.config(state=DISABLED)
            self.compare_btn.config(state=DISABLED)
            self.rollback_btn.config(state=DISABLED)
            self.branch_btn.config(state=DISABLED)
            self.export_diff_btn.config(state=DISABLED)
            
            self.submit_approval_btn.config(state=DISABLED)
            self.host_approve_btn.config(state=DISABLED)
            self.planner_approve_btn.config(state=DISABLED)
            self.customer_confirm_btn.config(state=DISABLED)
            self.export_approval_btn.config(state=DISABLED)
            self.complete_task_btn.config(state=DISABLED)
    
    def on_change_select(self, event):
        selection = self.change_tree.selection()
        if selection:
            self.delete_change_btn.config(state=NORMAL)
        else:
            self.delete_change_btn.config(state=DISABLED)
    
    def on_version_select(self, event):
        selection = self.version_tree.selection()
        if selection:
            item = self.version_tree.item(selection[0])
            version_id = item['values'][0]
            version = db_operations.get_version_by_id(version_id)
            if version:
                self.version_preview.config(state=tk.NORMAL)
                self.version_preview.delete('1.0', tk.END)
                self.version_preview.insert('1.0', version[4] if version[4] else '')
                self.version_preview.config(state=tk.DISABLED)
    
    def load_script_detail(self, script_id):
        script = db_operations.get_host_script_by_id(script_id)
        if script:
            self.no_data_label.grid_remove()
            keys = ['record_no', 'bride_name', 'groom_name', 'wedding_date', 'host_name', 
                    'current_version', 'feedback_round', 'finalized_status', 'remarks']
            for i, key in enumerate(keys):
                self.detail_vars[key].set(str(script[i + 1]))
            
            self.content_text.config(state=tk.NORMAL)
            self.content_text.delete('1.0', tk.END)
            self.content_text.insert('1.0', script[9] if script[9] else '')
            self.content_text.config(state=tk.DISABLED)
    
    def load_change_records(self, script_id):
        for item in self.change_tree.get_children():
            self.change_tree.delete(item)
        
        records = db_operations.get_change_records_by_script_id(script_id)
        for record in records:
            adopted = '是' if record[6] == 1 else '否'
            self.change_tree.insert('', END, values=(record[0], record[2], record[3], 
                                                      record[4], record[5], adopted))
    
    def load_version_history(self, script_id):
        for item in self.version_tree.get_children():
            self.version_tree.delete(item)
        
        versions = db_operations.get_version_history_by_script_id(script_id)
        for version in versions:
            self.version_tree.insert('', END, values=(version[0], version[2], version[3], 
                                                       version[5], version[7]))
    
    def load_rollback_records(self, script_id):
        for item in self.rollback_tree.get_children():
            self.rollback_tree.delete(item)
        
        records = db_operations.get_rollback_records_by_script_id(script_id)
        for record in records:
            self.rollback_tree.insert('', END, values=(record[0], record[2], record[3], 
                                                        record[4], record[5], record[6]))
    
    def search_scripts(self):
        keyword = self.search_var.get().strip()
        if not keyword:
            self.load_script_list()
            return
        
        scripts = db_operations.search_host_scripts(keyword)
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for script in scripts:
            self.tree.insert('', END, values=(script[0], script[1], script[2], script[3], script[4], 
                                              script[5], script[6], script[7], script[8]))
    
    def filter_scripts(self):
        host_name = self.host_combobox.get() if self.host_combobox.get() else None
        status = self.status_combobox.get() if self.status_combobox.get() else None
        
        scripts = db_operations.filter_host_scripts(host_name, status)
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        for script in scripts:
            self.tree.insert('', END, values=(script[0], script[1], script[2], script[3], script[4], 
                                              script[5], script[6], script[7], script[8]))
    
    def clear_filter(self):
        self.host_combobox.set('')
        self.status_combobox.set('')
        self.search_var.set('')
        self.load_script_list()
    
    def open_add_dialog(self):
        dialog = ScriptDialog(self.root, self)
        self.root.wait_window(dialog.top)
    
    def open_edit_dialog(self):
        if self.current_script_id:
            dialog = ScriptDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def delete_script(self):
        if self.current_script_id:
            if messagebox.askyesno("确认删除", "确定要删除这条主持词记录吗？"):
                success, error = db_operations.delete_host_script(self.current_script_id)
                if success:
                    messagebox.showinfo("成功", "删除成功")
                    self.load_script_list()
                    self.refresh_stats()
                    self.current_script_id = None
                    self.edit_btn.config(state=DISABLED)
                    self.delete_btn.config(state=DISABLED)
                    self.add_change_btn.config(state=DISABLED)
                    self.compare_btn.config(state=DISABLED)
                    self.rollback_btn.config(state=DISABLED)
                    self.branch_btn.config(state=DISABLED)
                    self.export_diff_btn.config(state=DISABLED)
                    self.no_data_label.grid()
                    self.content_text.config(state=tk.NORMAL)
                    self.content_text.delete('1.0', tk.END)
                    self.content_text.config(state=tk.DISABLED)
                else:
                    messagebox.showerror("错误", f"删除失败: {error}")
    
    def open_add_change_dialog(self):
        if self.current_script_id:
            script = db_operations.get_host_script_by_id(self.current_script_id)
            if script and script[8] == '已定稿':
                messagebox.showwarning("警告", "该主持词已定稿，无法新增变更记录")
                return
            
            dialog = ChangeRecordDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def delete_selected_change(self):
        selection = self.change_tree.selection()
        if selection:
            item = self.change_tree.item(selection[0])
            record_id = item['values'][0]
            
            if messagebox.askyesno("确认删除", "确定要删除这条变更记录吗？"):
                success, error = db_operations.delete_change_record(record_id)
                if success:
                    messagebox.showinfo("成功", "删除成功")
                    self.load_change_records(self.current_script_id)
                    self.load_script_list()
                    self.refresh_stats()
                else:
                    messagebox.showerror("错误", f"删除失败: {error}")
    
    def open_compare_dialog(self):
        if self.current_script_id:
            dialog = VersionCompareDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def open_rollback_dialog(self):
        if self.current_script_id:
            dialog = RollbackDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def open_branch_dialog(self):
        if self.current_script_id:
            dialog = BranchDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def export_diff_summary(self):
        if self.current_script_id:
            dialog = ExportDiffDialog(self.root, self, self.current_script_id)
            self.root.wait_window(dialog.top)
    
    def export_report(self):
        wb = openpyxl.Workbook()
        
        ws1 = wb.active
        ws1.title = "主持词列表"
        headers1 = ['记录编号', '新娘姓名', '新郎姓名', '婚礼日期', '主持人', 
                    '当前版本', '反馈轮次', '定稿状态', '备注', '更新时间']
        for i, header in enumerate(headers1):
            cell = ws1.cell(row=1, column=i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font.color = openpyxl.styles.colors.WHITE
        
        scripts = db_operations.get_all_host_scripts()
        for row_idx, script in enumerate(scripts, start=2):
            for col_idx, value in enumerate(script[1:11], start=1):
                ws1.cell(row=row_idx, column=col_idx, value=value)
        
        ws2 = wb.create_sheet(title="变更记录")
        headers2 = ['主持词编号', '修改日期', '修改段落', '修改原因', '反馈来源', '是否采纳']
        for i, header in enumerate(headers2):
            cell = ws2.cell(row=1, column=i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font.color = openpyxl.styles.colors.WHITE
        
        all_changes = []
        for script in scripts:
            changes = db_operations.get_change_records_by_script_id(script[0])
            for change in changes:
                all_changes.append((script[1], change[2], change[3], change[4], change[5], '是' if change[6] == 1 else '否'))
        
        for row_idx, change in enumerate(all_changes, start=2):
            for col_idx, value in enumerate(change, start=1):
                ws2.cell(row=row_idx, column=col_idx, value=value)
        
        ws3 = wb.create_sheet(title="高频问题")
        headers3 = ['问题描述', '出现次数', '首次出现', '最近出现']
        for i, header in enumerate(headers3):
            cell = ws3.cell(row=1, column=i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font.color = openpyxl.styles.colors.WHITE
        
        high_freq = db_operations.get_high_freq_issues()
        high_freq = [i for i in high_freq if i[1] >= 3]
        for row_idx, issue in enumerate(high_freq, start=2):
            for col_idx, value in enumerate(issue, start=1):
                ws3.cell(row=row_idx, column=col_idx, value=value)
        
        today = datetime.datetime.now().strftime('%Y%m%d')
        filename = f'主持词报表_{today}.xlsx'
        wb.save(filename)
        messagebox.showinfo("成功", f"报表已导出为: {filename}")

class ScriptDialog:
    def __init__(self, parent, app, script_id=None):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("编辑主持词" if script_id else "新增主持词")
        self.top.geometry("700x600")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="主持词信息")
        frame.pack(fill=X, padx=20, pady=10)
        
        fields = [
            ('新娘姓名', 'bride_name', True),
            ('新郎姓名', 'groom_name', True),
            ('婚礼日期', 'wedding_date', True),
            ('主持人', 'host_name', True),
            ('当前版本', 'current_version', False),
            ('反馈轮次', 'feedback_round', False),
            ('定稿状态', 'finalized_status', False),
            ('备注', 'remarks', False)
        ]
        
        self.entries = {}
        self.vars = {}
        
        for i, (label, key, required) in enumerate(fields):
            ttk.Label(frame, text=f"{label}{'*' if required else ''}:").grid(row=i, column=0, padx=10, pady=8, sticky=W)
            
            if key == 'current_version':
                var = tk.StringVar(value='1.0')
                entry = ttk.Entry(frame, textvariable=var, width=30)
            elif key == 'feedback_round':
                var = tk.StringVar(value='0')
                entry = ttk.Entry(frame, textvariable=var, width=30)
            elif key == 'finalized_status':
                var = tk.StringVar(value='未定稿')
                entry = ttk.Combobox(frame, textvariable=var, values=['未定稿', '已定稿'], width=27)
            elif key == 'remarks':
                var = tk.StringVar()
                entry = ttk.Text(frame, width=30, height=3)
                entry.grid(row=i, column=1, padx=10, pady=8)
                self.entries[key] = entry
                self.vars[key] = var
                continue
            else:
                var = tk.StringVar()
                entry = ttk.Entry(frame, textvariable=var, width=30)
            
            entry.grid(row=i, column=1, padx=10, pady=8)
            self.entries[key] = entry
            self.vars[key] = var
        
        content_frame = ttk.LabelFrame(self.top, text="主持词正文")
        content_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)
        
        self.content_text = tk.Text(content_frame, wrap=tk.WORD, font=('Microsoft YaHei', 11))
        self.content_text.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        if script_id:
            script = db_operations.get_host_script_by_id(script_id)
            if script:
                keys = ['bride_name', 'groom_name', 'wedding_date', 'host_name', 
                        'current_version', 'feedback_round', 'finalized_status']
                for i, key in enumerate(keys):
                    self.vars[key].set(str(script[i + 2]))
                self.entries['remarks'].insert('1.0', script[10] if script[10] else '')
                self.content_text.insert('1.0', script[9] if script[9] else '')
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="保存", command=self.save).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def save(self):
        bride_name = self.vars['bride_name'].get().strip()
        groom_name = self.vars['groom_name'].get().strip()
        wedding_date = self.vars['wedding_date'].get().strip()
        host_name = self.vars['host_name'].get().strip()
        
        try:
            current_version = float(self.vars['current_version'].get().strip())
        except ValueError:
            messagebox.showerror("错误", "版本号必须为数字")
            return
        
        try:
            feedback_round = int(self.vars['feedback_round'].get().strip())
        except ValueError:
            messagebox.showerror("错误", "反馈轮次必须为整数")
            return
        
        finalized_status = self.vars['finalized_status'].get()
        remarks = self.entries['remarks'].get('1.0', tk.END).strip()
        script_content = self.content_text.get('1.0', tk.END).strip()
        
        if not bride_name or not groom_name or not wedding_date or not host_name:
            messagebox.showerror("错误", "请填写必填字段")
            return
        
        if self.script_id:
            original_script = db_operations.get_host_script_by_id(self.script_id)
            if original_script and original_script[8] == '已定稿' and finalized_status != '已定稿':
                messagebox.showerror("错误", "已定稿的主持词不能修改为未定稿")
                return
            
            if current_version < original_script[6]:
                messagebox.showerror("错误", "版本号必须大于等于当前版本")
                return
            
            if feedback_round > 0:
                change_count = db_operations.get_change_record_count(self.script_id)
                if change_count == 0:
                    messagebox.showerror("错误", "反馈轮次大于0时，必须至少有一条变更记录")
                    return
            
            if db_operations.check_duplicate(bride_name, groom_name, wedding_date, self.script_id):
                messagebox.showerror("错误", "同一新人姓名与婚礼日期组合只能创建一份主持词")
                return
            
            success, error = db_operations.update_host_script(
                self.script_id, bride_name, groom_name, wedding_date, host_name,
                current_version, feedback_round, finalized_status, script_content, remarks
            )
        else:
            if db_operations.check_duplicate(bride_name, groom_name, wedding_date):
                messagebox.showerror("错误", "同一新人姓名与婚礼日期组合只能创建一份主持词")
                return
            
            success, error = db_operations.add_host_script(
                bride_name, groom_name, wedding_date, host_name,
                current_version, feedback_round, finalized_status, script_content, remarks
            )
        
        if success:
            messagebox.showinfo("成功", "保存成功")
            self.app.load_script_list()
            self.app.load_host_names()
            self.app.refresh_stats()
            if self.script_id:
                self.app.load_script_detail(self.script_id)
                self.app.load_version_history(self.script_id)
            self.top.destroy()
        else:
            messagebox.showerror("错误", f"保存失败: {error}")

class ChangeRecordDialog:
    def __init__(self, parent, app, script_id):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("新增变更记录")
        self.top.geometry("500x400")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="变更记录信息")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text="修改段落*:").grid(row=0, column=0, padx=10, pady=10, sticky=W)
        self.paragraph_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.paragraph_var, width=40).grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(frame, text="修改原因*:").grid(row=1, column=0, padx=10, pady=10, sticky=W)
        self.reason_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.reason_var, width=40).grid(row=1, column=1, padx=10, pady=10)
        
        ttk.Label(frame, text="反馈来源:").grid(row=2, column=0, padx=10, pady=10, sticky=W)
        self.source_var = tk.StringVar()
        ttk.Combobox(frame, textvariable=self.source_var, values=['客户', '策划师', '主持人', '其他'], width=37).grid(row=2, column=1, padx=10, pady=10)
        
        ttk.Label(frame, text="是否采纳:").grid(row=3, column=0, padx=10, pady=10, sticky=W)
        self.adopted_var = tk.IntVar(value=1)
        ttk.Radiobutton(frame, text="是", variable=self.adopted_var, value=1).grid(row=3, column=1, padx=10, pady=10, sticky=W)
        ttk.Radiobutton(frame, text="否", variable=self.adopted_var, value=0).grid(row=3, column=1, padx=60, pady=10, sticky=W)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="保存", command=self.save).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def save(self):
        modify_paragraph = self.paragraph_var.get().strip()
        modify_reason = self.reason_var.get().strip()
        feedback_source = self.source_var.get().strip()
        is_adopted = self.adopted_var.get()
        
        if not modify_paragraph or not modify_reason:
            messagebox.showerror("错误", "请填写必填字段")
            return
        
        success, error = db_operations.add_change_record(
            self.script_id, modify_paragraph, modify_reason, feedback_source, is_adopted
        )
        
        if success:
            script = db_operations.get_host_script_by_id(self.script_id)
            if script:
                new_version = round(script[6] + 0.1, 1)
                new_round = script[7] + 1
                db_operations.update_host_script(
                    self.script_id, script[2], script[3], script[4], script[5],
                    new_version, new_round, script[8], script[9], script[10]
                )
            
            messagebox.showinfo("成功", "保存成功")
            self.app.load_script_list()
            self.app.load_change_records(self.script_id)
            self.app.load_version_history(self.script_id)
            self.app.refresh_stats()
            self.top.destroy()
        else:
            messagebox.showerror("错误", f"保存失败: {error}")

class VersionCompareDialog:
    def __init__(self, parent, app, script_id):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("版本差异对比")
        self.top.geometry("1200x700")
        self.top.resizable(True, True)
        self.top.grab_set()
        
        self.versions = db_operations.get_version_history_by_script_id(script_id)
        if len(self.versions) < 2:
            messagebox.showwarning("警告", "至少需要两个版本才能进行对比")
            self.top.destroy()
            return
        
        select_frame = ttk.LabelFrame(self.top, text="选择版本")
        select_frame.pack(fill=X, padx=20, pady=10)
        
        ttk.Label(select_frame, text="版本A:").grid(row=0, column=0, padx=10, pady=10)
        self.version_a_var = tk.StringVar()
        version_a_values = [f"{v[2]} - {v[3]}" for v in self.versions]
        self.version_a_combobox = ttk.Combobox(select_frame, textvariable=self.version_a_var, values=version_a_values, width=30)
        self.version_a_combobox.grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(select_frame, text="版本B:").grid(row=0, column=2, padx=10, pady=10)
        self.version_b_var = tk.StringVar()
        self.version_b_combobox = ttk.Combobox(select_frame, textvariable=self.version_b_var, values=version_a_values, width=30)
        self.version_b_combobox.grid(row=0, column=3, padx=10, pady=10)
        
        quick_btn_frame = ttk.Frame(select_frame)
        quick_btn_frame.grid(row=0, column=4, padx=20)
        ttk.Button(quick_btn_frame, text="初稿 vs 定稿", command=self.select_first_last).pack(side=LEFT)
        
        ttk.Button(select_frame, text="开始对比", command=self.compare).grid(row=0, column=5, padx=20)
        
        diff_frame = ttk.LabelFrame(self.top, text="差异结果")
        diff_frame.pack(fill=BOTH, expand=True, padx=20, pady=10)
        
        self.diff_text = tk.Text(diff_frame, wrap=tk.WORD, font=('Consolas', 10))
        self.diff_text.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        scrollbar = ttk.Scrollbar(diff_frame, orient=VERTICAL, command=self.diff_text.yview)
        self.diff_text.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)
        
        self.diff_text.tag_configure('added', foreground='green', background='#d4edda')
        self.diff_text.tag_configure('removed', foreground='red', background='#f8d7da')
        self.diff_text.tag_configure('header', font=('Microsoft YaHei', 11, 'bold'))
        
        if len(self.versions) >= 2:
            self.version_a_combobox.current(0)
            self.version_b_combobox.current(-1)
            self.compare()
    
    def select_first_last(self):
        self.version_a_combobox.current(0)
        self.version_b_combobox.current(-1)
        self.compare()
    
    def compare(self):
        a_idx = self.version_a_combobox.current()
        b_idx = self.version_b_combobox.current()
        
        if a_idx < 0 or b_idx < 0:
            messagebox.showwarning("警告", "请选择两个版本")
            return
        
        version_a = self.versions[a_idx]
        version_b = self.versions[b_idx]
        
        content_a = version_a[4] if version_a[4] else ''
        content_b = version_b[4] if version_b[4] else ''
        
        lines_a = content_a.split('\n')
        lines_b = content_b.split('\n')
        
        differ = difflib.Differ()
        diff_result = differ.compare(lines_a, lines_b)
        
        self.diff_text.delete('1.0', tk.END)
        self.diff_text.insert('1.0', f"版本 {version_a[2]} ({version_a[3]}) → 版本 {version_b[2]} ({version_b[3]})\n\n", 'header')
        
        added_count = 0
        removed_count = 0
        
        for line in diff_result:
            if line.startswith('+ '):
                self.diff_text.insert(tk.END, line + '\n', 'added')
                added_count += 1
            elif line.startswith('- '):
                self.diff_text.insert(tk.END, line + '\n', 'removed')
                removed_count += 1
            elif line.startswith('? '):
                continue
            else:
                self.diff_text.insert(tk.END, line + '\n')
        
        self.diff_text.insert(tk.END, f"\n\n差异统计:\n", 'header')
        self.diff_text.insert(tk.END, f"新增段落: {added_count}\n")
        self.diff_text.insert(tk.END, f"删除段落: {removed_count}\n")

class RollbackDialog:
    def __init__(self, parent, app, script_id):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("版本回滚")
        self.top.geometry("500x400")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="选择目标版本")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text="目标版本:").pack(anchor=W, padx=10, pady=10)
        
        self.versions = db_operations.get_version_history_by_script_id(script_id)
        version_values = [f"{v[2]} - {v[3]} ({v[7]})" for v in self.versions]
        self.version_var = tk.StringVar()
        self.version_combobox = ttk.Combobox(frame, textvariable=self.version_var, values=version_values, width=50)
        self.version_combobox.pack(fill=X, padx=10, pady=10)
        
        ttk.Label(frame, text="回滚原因*:").pack(anchor=W, padx=10, pady=10)
        self.reason_text = tk.Text(frame, width=50, height=5)
        self.reason_text.pack(fill=X, padx=10, pady=10)
        
        ttk.Label(frame, text="操作人:").pack(anchor=W, padx=10, pady=10)
        self.operator_var = tk.StringVar(value='系统')
        ttk.Entry(frame, textvariable=self.operator_var, width=50).pack(fill=X, padx=10, pady=10)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="确认回滚", command=self.rollback).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def rollback(self):
        idx = self.version_combobox.current()
        if idx < 0:
            messagebox.showwarning("警告", "请选择目标版本")
            return
        
        rollback_reason = self.reason_text.get('1.0', tk.END).strip()
        if not rollback_reason:
            messagebox.showerror("错误", "请填写回滚原因")
            return
        
        operated_by = self.operator_var.get().strip() or '系统'
        
        target_version = self.versions[idx]
        target_version_number = target_version[2]
        
        script = db_operations.get_host_script_by_id(self.script_id)
        if script and script[6] == target_version_number:
            messagebox.showwarning("警告", "当前版本已是目标版本，无需回滚")
            return
        
        if messagebox.askyesno("确认回滚", f"确定要将当前版本回滚到版本 {target_version_number} 吗？\n回滚原因: {rollback_reason}"):
            success, error = db_operations.rollback_to_version(self.script_id, target_version_number, rollback_reason, operated_by)
            if success:
                messagebox.showinfo("成功", "回滚成功")
                self.app.load_script_list()
                self.app.load_script_detail(self.script_id)
                self.app.load_version_history(self.script_id)
                self.app.load_rollback_records(self.script_id)
                self.top.destroy()
            else:
                messagebox.showerror("错误", f"回滚失败: {error}")

class BranchDialog:
    def __init__(self, parent, app, script_id):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("复制为新版本分支")
        self.top.geometry("500x350")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="新版本信息")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        script = db_operations.get_host_script_by_id(script_id)
        current_version = script[6] if script else 1.0
        new_version = round(current_version + 0.1, 1)
        
        ttk.Label(frame, text="源版本:").grid(row=0, column=0, padx=10, pady=10, sticky=W)
        ttk.Label(frame, text=str(current_version)).grid(row=0, column=1, padx=10, pady=10, sticky=W)
        
        ttk.Label(frame, text="新版本号*:").grid(row=1, column=0, padx=10, pady=10, sticky=W)
        self.new_version_var = tk.StringVar(value=str(new_version))
        ttk.Entry(frame, textvariable=self.new_version_var, width=20).grid(row=1, column=1, padx=10, pady=10)
        
        ttk.Label(frame, text="版本标签:").grid(row=2, column=0, padx=10, pady=10, sticky=W)
        self.label_var = tk.StringVar(value=f'分支{new_version}')
        ttk.Entry(frame, textvariable=self.label_var, width=40).grid(row=2, column=1, padx=10, pady=10)
        
        info_label = ttk.Label(frame, text="注意：复制后将创建新版本分支，原定稿版本将被保留不可修改", 
                               foreground='red', wraplength=400)
        info_label.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="确认创建", command=self.create_branch).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def create_branch(self):
        try:
            new_version = float(self.new_version_var.get().strip())
        except ValueError:
            messagebox.showerror("错误", "版本号必须为数字")
            return
        
        script = db_operations.get_host_script_by_id(self.script_id)
        if not script:
            messagebox.showerror("错误", "主持词不存在")
            return
        
        current_version = script[6]
        if new_version <= current_version:
            messagebox.showerror("错误", "新版本号必须大于当前版本号")
            return
        
        label = self.label_var.get().strip()
        
        if messagebox.askyesno("确认创建分支", f"确定要基于版本 {current_version} 创建新版本 {new_version} 吗？"):
            success, error = db_operations.create_version_branch(self.script_id, current_version, new_version, label)
            if success:
                messagebox.showinfo("成功", f"新版本分支 {new_version} 创建成功")
                self.app.load_script_list()
                self.app.load_script_detail(self.script_id)
                self.app.load_version_history(self.script_id)
                self.top.destroy()
            else:
                messagebox.showerror("错误", f"创建失败: {error}")

class ExportDiffDialog:
    def __init__(self, parent, app, script_id):
        self.app = app
        self.script_id = script_id
        
        self.top = ttk.Toplevel(parent)
        self.top.title("导出差异摘要")
        self.top.geometry("500x350")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="选择对比版本")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        self.versions = db_operations.get_version_history_by_script_id(script_id)
        if len(self.versions) < 2:
            messagebox.showwarning("警告", "至少需要两个版本才能导出差异摘要")
            self.top.destroy()
            return
        
        ttk.Label(frame, text="版本A:").grid(row=0, column=0, padx=10, pady=10)
        self.version_a_var = tk.StringVar()
        version_values = [f"{v[2]} - {v[3]}" for v in self.versions]
        self.version_a_combobox = ttk.Combobox(frame, textvariable=self.version_a_var, values=version_values, width=30)
        self.version_a_combobox.grid(row=0, column=1, padx=10, pady=10)
        
        ttk.Label(frame, text="版本B:").grid(row=1, column=0, padx=10, pady=10)
        self.version_b_var = tk.StringVar()
        self.version_b_combobox = ttk.Combobox(frame, textvariable=self.version_b_var, values=version_values, width=30)
        self.version_b_combobox.grid(row=1, column=1, padx=10, pady=10)
        
        quick_btn_frame = ttk.Frame(frame)
        quick_btn_frame.grid(row=2, column=0, columnspan=2, pady=10)
        ttk.Button(quick_btn_frame, text="初稿 vs 定稿", command=self.select_first_last).pack(side=LEFT)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="导出到Excel", command=self.export).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
        
        if len(self.versions) >= 2:
            self.version_a_combobox.current(0)
            self.version_b_combobox.current(-1)
    
    def select_first_last(self):
        self.version_a_combobox.current(0)
        self.version_b_combobox.current(-1)
    
    def export(self):
        a_idx = self.version_a_combobox.current()
        b_idx = self.version_b_combobox.current()
        
        if a_idx < 0 or b_idx < 0:
            messagebox.showwarning("警告", "请选择两个版本")
            return
        
        version_a = self.versions[a_idx]
        version_b = self.versions[b_idx]
        
        content_a = version_a[4] if version_a[4] else ''
        content_b = version_b[4] if version_b[4] else ''
        
        lines_a = content_a.split('\n')
        lines_b = content_b.split('\n')
        
        differ = difflib.Differ()
        diff_result = list(differ.compare(lines_a, lines_b))
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "差异摘要"
        
        title_row = 1
        ws.merge_cells(f'A{title_row}:C{title_row}')
        title_cell = ws.cell(row=title_row, column=1, value=f"主持词差异对比报告")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        info_row = 2
        ws.cell(row=info_row, column=1, value=f"对比版本: {version_a[2]} ({version_a[3]}) → {version_b[2]} ({version_b[3]})")
        
        ws.cell(row=3, column=1, value="脚本ID")
        ws.cell(row=3, column=2, value=self.script_id)
        ws.cell(row=4, column=1, value="导出时间")
        ws.cell(row=4, column=2, value=datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        
        header_row = 6
        headers = ['操作类型', '内容', '说明']
        for i, header in enumerate(headers):
            cell = ws.cell(row=header_row, column=i + 1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.font.color = openpyxl.styles.colors.WHITE
            cell.alignment = Alignment(horizontal='center')
        
        data_row = header_row + 1
        added_count = 0
        removed_count = 0
        
        for line in diff_result:
            if line.startswith('+ '):
                ws.cell(row=data_row, column=1, value='新增')
                ws.cell(row=data_row, column=2, value=line[2:])
                ws.cell(row=data_row, column=3, value=f'版本{version_b[2]}新增内容')
                ws.cell(row=data_row, column=2).fill = PatternFill(start_color='d4edda', end_color='d4edda', fill_type='solid')
                added_count += 1
                data_row += 1
            elif line.startswith('- '):
                ws.cell(row=data_row, column=1, value='删除')
                ws.cell(row=data_row, column=2, value=line[2:])
                ws.cell(row=data_row, column=3, value=f'版本{version_a[2]}删除内容')
                ws.cell(row=data_row, column=2).fill = PatternFill(start_color='f8d7da', end_color='f8d7da', fill_type='solid')
                removed_count += 1
                data_row += 1
            elif line.startswith('? '):
                continue
            else:
                ws.cell(row=data_row, column=1, value='保留')
                ws.cell(row=data_row, column=2, value=line[2:])
                ws.cell(row=data_row, column=3, value='两版本相同')
                data_row += 1
        
        stats_row = data_row + 2
        ws.cell(row=stats_row, column=1, value="差异统计")
        ws.cell(row=stats_row + 1, column=1, value="新增段落数")
        ws.cell(row=stats_row + 1, column=2, value=added_count)
        ws.cell(row=stats_row + 2, column=1, value="删除段落数")
        ws.cell(row=stats_row + 2, column=2, value=removed_count)
        ws.cell(row=stats_row + 3, column=1, value="总变更数")
        ws.cell(row=stats_row + 3, column=2, value=added_count + removed_count)
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 40
        
        today = datetime.datetime.now().strftime('%Y%m%d')
        filename = f'差异摘要_{version_a[2]}_vs_{version_b[2]}_{today}.xlsx'
        wb.save(filename)
        
        messagebox.showinfo("成功", f"差异摘要已导出为: {filename}")
        self.top.destroy()

class SubmitApprovalDialog:
    def __init__(self, parent, app, script_id, version_number):
        self.app = app
        self.script_id = script_id
        self.version_number = version_number
        
        self.top = ttk.Toplevel(parent)
        self.top.title("提交审批")
        self.top.geometry("500x350")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="提交审批信息")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text=f"脚本ID: {script_id}").pack(anchor=W, padx=10, pady=5)
        ttk.Label(frame, text=f"版本号: {version_number}").pack(anchor=W, padx=10, pady=5)
        
        ttk.Label(frame, text="操作人角色*:").pack(anchor=W, padx=10, pady=10)
        self.role_var = tk.StringVar(value='主持人')
        self.role_combobox = ttk.Combobox(frame, textvariable=self.role_var, values=['主持人', '策划师'], width=45)
        self.role_combobox.pack(fill=X, padx=10, pady=5)
        
        ttk.Label(frame, text="操作人姓名:").pack(anchor=W, padx=10, pady=10)
        self.operator_var = tk.StringVar(value='系统')
        ttk.Entry(frame, textvariable=self.operator_var, width=50).pack(fill=X, padx=10, pady=5)
        
        ttk.Label(frame, text="备注:").pack(anchor=W, padx=10, pady=10)
        self.remark_text = tk.Text(frame, width=50, height=4)
        self.remark_text.pack(fill=X, padx=10, pady=5)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="确认提交", command=self.submit).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def submit(self):
        operator_role = self.role_var.get()
        operator_name = self.operator_var.get().strip() or '系统'
        remark = self.remark_text.get('1.0', tk.END).strip()
        
        if messagebox.askyesno("确认提交", f"确定要提交版本 {self.version_number} 进行审批吗？\n操作人: {operator_name} ({operator_role})"):
            flow_id, error = db_operations.submit_for_approval(self.script_id, self.version_number, operator_name, operator_role)
            if flow_id:
                messagebox.showinfo("成功", "审批已提交")
                self.app.load_approval_info(self.script_id)
                self.app.load_script_list()
                self.top.destroy()
            else:
                messagebox.showerror("错误", f"提交失败: {error}")

class ApproveDialog:
    def __init__(self, parent, app, script_id, version_number, role):
        self.app = app
        self.script_id = script_id
        self.version_number = version_number
        self.role = role
        
        self.top = ttk.Toplevel(parent)
        self.top.title(f"{role}确认")
        self.top.geometry("500x400")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text=f"{role}确认信息")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text=f"脚本ID: {script_id}").pack(anchor=W, padx=10, pady=5)
        ttk.Label(frame, text=f"版本号: {version_number}").pack(anchor=W, padx=10, pady=5)
        ttk.Label(frame, text=f"当前角色: {role}").pack(anchor=W, padx=10, pady=5)
        
        ttk.Label(frame, text="操作人姓名:").pack(anchor=W, padx=10, pady=10)
        self.operator_var = tk.StringVar(value='系统')
        ttk.Entry(frame, textvariable=self.operator_var, width=50).pack(fill=X, padx=10, pady=5)
        
        ttk.Label(frame, text="操作类型:").pack(anchor=W, padx=10, pady=10)
        self.action_var = tk.StringVar(value='确认')
        ttk.Radiobutton(frame, text="确认", variable=self.action_var, value='确认').pack(anchor=W, padx=20, pady=5)
        ttk.Radiobutton(frame, text="退回", variable=self.action_var, value='退回').pack(anchor=W, padx=20, pady=5)
        ttk.Radiobutton(frame, text="催办", variable=self.action_var, value='催办').pack(anchor=W, padx=20, pady=5)
        
        ttk.Label(frame, text="备注/退回原因:").pack(anchor=W, padx=10, pady=10)
        self.remark_text = tk.Text(frame, width=50, height=4)
        self.remark_text.pack(fill=X, padx=10, pady=5)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="确认操作", command=self.approve).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def approve(self):
        action = self.action_var.get()
        operator_name = self.operator_var.get().strip() or '系统'
        remark = self.remark_text.get('1.0', tk.END).strip()
        
        flow = db_operations.get_approval_flow_by_script_and_version(self.script_id, self.version_number)
        if not flow:
            messagebox.showerror("错误", "审批流程不存在，请先提交审批")
            return
        
        flow_id = flow[0]
        
        if messagebox.askyesno("确认操作", f"确定要执行【{action}】操作吗？\n角色: {self.role}\n操作人: {operator_name}"):
            success, error = db_operations.approve_by_role(flow_id, self.script_id, self.version_number, action, self.role, operator_name, remark)
            if success:
                messagebox.showinfo("成功", f"{self.role}{action}成功")
                self.app.load_approval_info(self.script_id)
                self.app.load_script_list()
                self.top.destroy()
            else:
                messagebox.showerror("错误", f"操作失败: {error}")

class CustomerConfirmDialog:
    def __init__(self, parent, app, script_id, version_number):
        self.app = app
        self.script_id = script_id
        self.version_number = version_number
        
        self.top = ttk.Toplevel(parent)
        self.top.title("客户确认")
        self.top.geometry("600x500")
        self.top.resizable(False, False)
        self.top.grab_set()
        
        frame = ttk.LabelFrame(self.top, text="客户确认信息")
        frame.pack(fill=BOTH, expand=True, padx=20, pady=20)
        
        ttk.Label(frame, text=f"脚本ID: {script_id}").pack(anchor=W, padx=10, pady=5)
        ttk.Label(frame, text=f"版本号: {version_number}").pack(anchor=W, padx=10, pady=5)
        
        ttk.Label(frame, text="确认结果*:").pack(anchor=W, padx=10, pady=10)
        self.result_var = tk.StringVar()
        ttk.Radiobutton(frame, text="通过", variable=self.result_var, value='通过').pack(anchor=W, padx=20, pady=5)
        ttk.Radiobutton(frame, text="退回", variable=self.result_var, value='退回').pack(anchor=W, padx=20, pady=5)
        
        ttk.Label(frame, text="确认人姓名*:").pack(anchor=W, padx=10, pady=10)
        self.confirmed_by_var = tk.StringVar()
        ttk.Entry(frame, textvariable=self.confirmed_by_var, width=50).pack(fill=X, padx=10, pady=5)
        
        ttk.Label(frame, text="反馈意见:").pack(anchor=W, padx=10, pady=10)
        self.feedback_text = tk.Text(frame, width=50, height=6)
        self.feedback_text.pack(fill=X, padx=10, pady=5)
        
        info_label = ttk.Label(frame, text="注意：选择'通过'将直接标记为正式定稿；选择'退回'将自动生成反馈任务", 
                               foreground='red', wraplength=500)
        info_label.pack(anchor=W, padx=10, pady=10)
        
        btn_frame = ttk.Frame(self.top)
        btn_frame.pack(fill=X, padx=20, pady=(0, 20))
        
        ttk.Button(btn_frame, text="确认提交", command=self.confirm).pack(side=LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.top.destroy).pack(side=RIGHT, padx=20)
    
    def confirm(self):
        confirm_result = self.result_var.get()
        confirmed_by = self.confirmed_by_var.get().strip()
        feedback = self.feedback_text.get('1.0', tk.END).strip()
        
        if not confirm_result:
            messagebox.showerror("错误", "请选择确认结果")
            return
        
        if not confirmed_by:
            messagebox.showerror("错误", "请填写确认人姓名")
            return
        
        action_text = "通过并定稿" if confirm_result == '通过' else "退回"
        
        if messagebox.askyesno("确认提交", f"确定要【{action_text}】吗？\n确认人: {confirmed_by}\n反馈意见: {feedback[:50]}..." if feedback else f"确定要【{action_text}】吗？\n确认人: {confirmed_by}"):
            success, error = db_operations.customer_confirm(self.script_id, self.version_number, confirm_result, feedback, confirmed_by)
            if success:
                messagebox.showinfo("成功", f"客户{action_text}成功")
                self.app.load_approval_info(self.script_id)
                self.app.load_script_list()
                self.app.load_script_detail(self.script_id)
                self.app.refresh_stats()
                self.top.destroy()
            else:
                messagebox.showerror("错误", f"操作失败: {error}")

if __name__ == "__main__":
    root = ttkb.Window(themename='cosmo')
    app = WeddingScriptApp(root)
    root.mainloop()
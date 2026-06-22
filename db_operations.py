from database import execute_query, execute_non_query
import datetime

def generate_record_no():
    today = datetime.datetime.now().strftime('%Y%m%d')
    count = execute_query('SELECT COUNT(*) FROM host_script WHERE record_no LIKE ?', (f'{today}%',))[0][0]
    return f'{today}{str(count + 1).zfill(3)}'

def add_host_script(bride_name, groom_name, wedding_date, host_name, current_version=1.0, feedback_round=0, finalized_status='未定稿', script_content='', remarks=''):
    record_no = generate_record_no()
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        script_id = execute_non_query('''
            INSERT INTO host_script (record_no, bride_name, groom_name, wedding_date, host_name, 
                                    current_version, feedback_round, finalized_status, script_content, remarks, 
                                    created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (record_no, bride_name, groom_name, wedding_date, host_name, current_version, feedback_round, finalized_status, script_content, remarks, now, now))
        save_version_history(script_id, current_version, '初稿', script_content, finalized_status)
        return script_id, None
    except Exception as e:
        return None, str(e)

def get_all_host_scripts():
    return execute_query('''
        SELECT id, record_no, bride_name, groom_name, wedding_date, host_name, 
               current_version, feedback_round, finalized_status, script_content, remarks, updated_at
        FROM host_script ORDER BY updated_at DESC
    ''')

def get_host_script_by_id(script_id):
    return execute_query('''
        SELECT id, record_no, bride_name, groom_name, wedding_date, host_name, 
               current_version, feedback_round, finalized_status, script_content, remarks
        FROM host_script WHERE id = ?
    ''', (script_id,), fetch_one=True)

def update_host_script(script_id, bride_name, groom_name, wedding_date, host_name, current_version, feedback_round, finalized_status, script_content='', remarks=''):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        existing_version = execute_query('SELECT status FROM version_history WHERE script_id = ? AND version_number = ?', 
                                        (script_id, current_version), fetch_one=True)
        if existing_version and existing_version[0] == '已定稿':
            return False, '已定稿版本不可篡改，请创建新版本分支'
        
        execute_non_query('''
            UPDATE host_script SET bride_name=?, groom_name=?, wedding_date=?, host_name=?, 
                                  current_version=?, feedback_round=?, finalized_status=?, 
                                  script_content=?, remarks=?, updated_at=?
            WHERE id = ?
        ''', (bride_name, groom_name, wedding_date, host_name, current_version, feedback_round, finalized_status, script_content, remarks, now, script_id))
        
        label = '定稿' if finalized_status == '已定稿' else '更新'
        save_version_history(script_id, current_version, label, script_content, finalized_status)
        return True, None
    except Exception as e:
        return False, str(e)

def delete_host_script(script_id):
    try:
        execute_non_query('DELETE FROM host_script WHERE id = ?', (script_id,))
        return True, None
    except Exception as e:
        return False, str(e)

def check_duplicate(bride_name, groom_name, wedding_date, exclude_id=None):
    if exclude_id:
        count = execute_query('''
            SELECT COUNT(*) FROM host_script 
            WHERE bride_name=? AND groom_name=? AND wedding_date=? AND id != ?
        ''', (bride_name, groom_name, wedding_date, exclude_id), fetch_one=True)[0]
    else:
        count = execute_query('''
            SELECT COUNT(*) FROM host_script 
            WHERE bride_name=? AND groom_name=? AND wedding_date=?
        ''', (bride_name, groom_name, wedding_date), fetch_one=True)[0]
    return count > 0

def add_change_record(script_id, modify_paragraph, modify_reason, feedback_source, is_adopted=1):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        record_id = execute_non_query('''
            INSERT INTO change_record (script_id, modify_date, modify_paragraph, 
                                       modify_reason, feedback_source, is_adopted, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (script_id, now, modify_paragraph, modify_reason, feedback_source, is_adopted, now))
        check_high_freq_issue(modify_reason)
        return record_id, None
    except Exception as e:
        return None, str(e)

def get_change_records_by_script_id(script_id):
    return execute_query('''
        SELECT id, script_id, modify_date, modify_paragraph, modify_reason, 
               feedback_source, is_adopted, created_at
        FROM change_record WHERE script_id = ? ORDER BY created_at DESC
    ''', (script_id,))

def get_change_record_count(script_id):
    count = execute_query('SELECT COUNT(*) FROM change_record WHERE script_id = ?', (script_id,), fetch_one=True)[0]
    return count

def delete_change_record(record_id):
    try:
        execute_non_query('DELETE FROM change_record WHERE id = ?', (record_id,))
        return True, None
    except Exception as e:
        return False, str(e)

def check_high_freq_issue(reason_text):
    existing = execute_query('SELECT id, occurrence_count FROM high_freq_issues WHERE issue_text = ?', (reason_text,), fetch_one=True)
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    if existing:
        issue_id, count = existing
        new_count = count + 1
        execute_non_query('''
            UPDATE high_freq_issues SET occurrence_count=?, last_occurrence=? WHERE id = ?
        ''', (new_count, now, issue_id))
    else:
        execute_non_query('''
            INSERT INTO high_freq_issues (issue_text, occurrence_count, first_occurrence, last_occurrence)
            VALUES (?, 1, ?, ?)
        ''', (reason_text, now, now))

def get_high_freq_issues():
    return execute_query('''
        SELECT issue_text, occurrence_count, first_occurrence, last_occurrence
        FROM high_freq_issues ORDER BY occurrence_count DESC
    ''')

def search_host_scripts(keyword):
    return execute_query('''
        SELECT id, record_no, bride_name, groom_name, wedding_date, host_name, 
               current_version, feedback_round, finalized_status, script_content, remarks, updated_at
        FROM host_script 
        WHERE record_no LIKE ? OR bride_name LIKE ? OR groom_name LIKE ? OR 
              host_name LIKE ? OR remarks LIKE ? OR script_content LIKE ?
        ORDER BY updated_at DESC
    ''', (f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%', f'%{keyword}%'))

def filter_host_scripts(host_name=None, finalized_status=None, date_range=None):
    sql = '''
        SELECT id, record_no, bride_name, groom_name, wedding_date, host_name, 
               current_version, feedback_round, finalized_status, script_content, remarks, updated_at
        FROM host_script WHERE 1=1
    '''
    params = []
    
    if host_name:
        sql += ' AND host_name = ?'
        params.append(host_name)
    
    if finalized_status:
        sql += ' AND finalized_status = ?'
        params.append(finalized_status)
    
    if date_range and len(date_range) == 2:
        sql += ' AND wedding_date BETWEEN ? AND ?'
        params.extend(date_range)
    
    sql += ' ORDER BY updated_at DESC'
    return execute_query(sql, params)

def get_stats_overview():
    total = execute_query('SELECT COUNT(*) FROM host_script', fetch_one=True)[0]
    finalized = execute_query("SELECT COUNT(*) FROM host_script WHERE finalized_status = '已定稿'", fetch_one=True)[0]
    avg_round = execute_query("SELECT AVG(feedback_round) FROM host_script", fetch_one=True)[0]
    total_changes = execute_query('SELECT COUNT(*) FROM change_record', fetch_one=True)[0]
    return {
        'total': total,
        'finalized': finalized,
        'avg_round': round(avg_round, 1) if avg_round else 0,
        'total_changes': total_changes
    }

def get_version_stats():
    return execute_query('''
        SELECT current_version, COUNT(*) as cnt 
        FROM host_script GROUP BY current_version ORDER BY current_version
    ''')

def get_feedback_source_stats():
    return execute_query('''
        SELECT feedback_source, COUNT(*) as cnt 
        FROM change_record WHERE feedback_source IS NOT NULL 
        GROUP BY feedback_source ORDER BY cnt DESC
    ''')

def get_change_reason_stats():
    return execute_query('''
        SELECT modify_reason, COUNT(*) as cnt 
        FROM change_record GROUP BY modify_reason ORDER BY cnt DESC LIMIT 10
    ''')

def get_modify_paragraph_stats():
    return execute_query('''
        SELECT modify_paragraph, COUNT(*) as cnt 
        FROM change_record GROUP BY modify_paragraph ORDER BY cnt DESC
    ''')

def get_host_names():
    return execute_query('SELECT DISTINCT host_name FROM host_script ORDER BY host_name')

def save_version_history(script_id, version_number, version_label, script_content, status='草稿'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        existing = execute_query('SELECT id, status FROM version_history WHERE script_id = ? AND version_number = ?', 
                                (script_id, version_number), fetch_one=True)
        if existing:
            if existing[1] == '已定稿':
                return False, '已定稿版本不可篡改'
            execute_non_query('''
                UPDATE version_history SET version_label=?, script_content=?, status=?, created_at=?
                WHERE script_id = ? AND version_number = ?
            ''', (version_label, script_content, status, now, script_id, version_number))
        else:
            execute_non_query('''
                INSERT INTO version_history (script_id, version_number, version_label, script_content, status, created_at)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (script_id, version_number, version_label, script_content, status, now))
        return True, None
    except Exception as e:
        return False, str(e)

def get_version_history_by_script_id(script_id):
    return execute_query('''
        SELECT id, script_id, version_number, version_label, script_content, status, created_by, created_at
        FROM version_history WHERE script_id = ? ORDER BY version_number ASC
    ''', (script_id,))

def get_version_by_id(version_id):
    return execute_query('''
        SELECT id, script_id, version_number, version_label, script_content, status, created_by, created_at
        FROM version_history WHERE id = ?
    ''', (version_id,), fetch_one=True)

def get_version_by_script_id_and_version(script_id, version_number):
    return execute_query('''
        SELECT id, script_id, version_number, version_label, script_content, status, created_by, created_at
        FROM version_history WHERE script_id = ? AND version_number = ?
    ''', (script_id, version_number), fetch_one=True)

def create_version_branch(script_id, source_version, new_version_number, new_version_label=''):
    source = get_version_by_script_id_and_version(script_id, source_version)
    if not source:
        return False, '源版本不存在'
    
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        execute_non_query('''
            INSERT INTO version_history (script_id, version_number, version_label, script_content, status, created_by, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (script_id, new_version_number, new_version_label if new_version_label else f'分支{new_version_number}', 
              source[4], '草稿', '系统', now))
        
        script = get_host_script_by_id(script_id)
        if script:
            execute_non_query('''
                UPDATE host_script SET current_version=?, script_content=?, finalized_status='未定稿', updated_at=?
                WHERE id = ?
            ''', (new_version_number, source[4], now, script_id))
        
        return True, None
    except Exception as e:
        return False, str(e)

def rollback_to_version(script_id, to_version_number, rollback_reason, operated_by='系统'):
    target_version = get_version_by_script_id_and_version(script_id, to_version_number)
    if not target_version:
        return False, '目标版本不存在'
    
    script = get_host_script_by_id(script_id)
    if not script:
        return False, '主持词不存在'
    
    from_version = script[6]
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    try:
        execute_non_query('''
            INSERT INTO rollback_record (script_id, from_version, to_version, rollback_reason, operated_by, operated_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (script_id, from_version, to_version_number, rollback_reason, operated_by, now))
        
        execute_non_query('''
            UPDATE host_script SET current_version=?, script_content=?, updated_at=?
            WHERE id = ?
        ''', (to_version_number, target_version[4], now, script_id))
        
        save_version_history(script_id, to_version_number, '回滚恢复', target_version[4], script[8])
        
        return True, None
    except Exception as e:
        return False, str(e)

def get_rollback_records_by_script_id(script_id):
    return execute_query('''
        SELECT id, script_id, from_version, to_version, rollback_reason, operated_by, operated_at
        FROM rollback_record WHERE script_id = ? ORDER BY operated_at DESC
    ''', (script_id,))

def submit_for_approval(script_id, version_number, operator_name='系统', operator_role='主持人'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        existing = execute_query('''
            SELECT id FROM approval_flow WHERE script_id = ? AND version_number = ?
        ''', (script_id, version_number), fetch_one=True)
        
        if existing:
            execute_non_query('''
                UPDATE approval_flow SET status='待确认', current_approver='主持人', updated_at=?
                WHERE script_id = ? AND version_number = ?
            ''', (now, script_id, version_number))
            flow_id = existing[0]
        else:
            flow_id = execute_non_query('''
                INSERT INTO approval_flow (script_id, version_number, status, current_approver, created_at, updated_at)
                VALUES (?, ?, '待确认', '主持人', ?, ?)
            ''', (script_id, version_number, now, now))
        
        add_approval_history(flow_id, script_id, version_number, '提交审批', operator_role, operator_name, '提交版本审批')
        
        return flow_id, None
    except Exception as e:
        return None, str(e)

def update_approval_flow(flow_id, status=None, current_approver=None, customer_view_time=None,
                          customer_confirm_result=None, customer_feedback=None, customer_confirmed_by=None, 
                          customer_confirmed_at=None):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        sql = 'UPDATE approval_flow SET updated_at=?'
        params = [now]
        
        if status:
            sql += ', status=?'
            params.append(status)
        if current_approver:
            sql += ', current_approver=?'
            params.append(current_approver)
        if customer_view_time:
            sql += ', customer_view_time=?'
            params.append(customer_view_time)
        if customer_confirm_result:
            sql += ', customer_confirm_result=?'
            params.append(customer_confirm_result)
        if customer_feedback:
            sql += ', customer_feedback=?'
            params.append(customer_feedback)
        if customer_confirmed_by:
            sql += ', customer_confirmed_by=?'
            params.append(customer_confirmed_by)
        if customer_confirmed_at:
            sql += ', customer_confirmed_at=?'
            params.append(customer_confirmed_at)
        
        sql += ' WHERE id=?'
        params.append(flow_id)
        
        execute_non_query(sql, params)
        return True, None
    except Exception as e:
        return False, str(e)

def get_approval_flow_by_script_and_version(script_id, version_number):
    return execute_query('''
        SELECT id, script_id, version_number, status, current_approver, customer_view_time,
               customer_confirm_result, customer_feedback, customer_confirmed_by, customer_confirmed_at,
               created_at, updated_at
        FROM approval_flow WHERE script_id = ? AND version_number = ?
    ''', (script_id, version_number), fetch_one=True)

def get_approval_flow_by_script_id(script_id):
    return execute_query('''
        SELECT id, script_id, version_number, status, current_approver, customer_view_time,
               customer_confirm_result, customer_feedback, customer_confirmed_by, customer_confirmed_at,
               created_at, updated_at
        FROM approval_flow WHERE script_id = ? ORDER BY version_number DESC
    ''', (script_id,))

def get_approval_flow_by_id(flow_id):
    return execute_query('''
        SELECT id, script_id, version_number, status, current_approver, customer_view_time,
               customer_confirm_result, customer_feedback, customer_confirmed_by, customer_confirmed_at,
               created_at, updated_at
        FROM approval_flow WHERE id = ?
    ''', (flow_id,), fetch_one=True)

def record_customer_view(script_id, version_number):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        flow = get_approval_flow_by_script_and_version(script_id, version_number)
        if flow:
            return update_approval_flow(flow[0], customer_view_time=now)
        return False, '审批流程不存在'
    except Exception as e:
        return False, str(e)

def customer_confirm(script_id, version_number, confirm_result, feedback, confirmed_by, modify_paragraph='全文'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        flow = get_approval_flow_by_script_and_version(script_id, version_number)
        if not flow:
            return False, '审批流程不存在'
        
        flow_id = flow[0]
        
        if confirm_result == '通过':
            success, error = update_approval_flow(
                flow_id, 
                status='已通过',
                customer_confirm_result='通过',
                customer_feedback=feedback,
                customer_confirmed_by=confirmed_by,
                customer_confirmed_at=now
            )
            if success:
                execute_non_query('''
                    UPDATE host_script SET finalized_status='已定稿', updated_at=? WHERE id=?
                ''', (now, script_id))
                
                execute_non_query('''
                    UPDATE version_history SET status='已定稿', updated_at=? WHERE script_id=? AND version_number=?
                ''', (now, script_id, version_number))
                
                add_approval_history(flow_id, script_id, version_number, '客户确认通过', '客户', confirmed_by, feedback or '无意见')
        
        elif confirm_result == '退回':
            success, error = update_approval_flow(
                flow_id, 
                status='已退回',
                customer_confirm_result='退回',
                customer_feedback=feedback,
                customer_confirmed_by=confirmed_by,
                customer_confirmed_at=now
            )
            if success:
                add_approval_history(flow_id, script_id, version_number, '客户退回', '客户', confirmed_by, f"{modify_paragraph}: {feedback}" if feedback else modify_paragraph)
                
                create_feedback_task(flow_id, script_id, version_number, modify_paragraph, feedback)
        
        else:
            return False, '无效的确认结果'
        
        return success, error
    except Exception as e:
        return False, str(e)

def add_approval_history(flow_id, script_id, version_number, action, operator_role, operator_name='系统', remark=''):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        execute_non_query('''
            INSERT INTO approval_history (flow_id, script_id, version_number, action, operator_role, operator_name, remark, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (flow_id, script_id, version_number, action, operator_role, operator_name, remark, now))
        return True, None
    except Exception as e:
        return False, str(e)

def get_approval_history_by_flow_id(flow_id):
    return execute_query('''
        SELECT id, flow_id, script_id, version_number, action, operator_role, operator_name, remark, created_at
        FROM approval_history WHERE flow_id = ? ORDER BY created_at ASC
    ''', (flow_id,))

def get_approval_history_by_script_id(script_id):
    return execute_query('''
        SELECT ah.id, ah.flow_id, ah.script_id, ah.version_number, ah.action, ah.operator_role, 
               ah.operator_name, ah.remark, ah.created_at, af.status
        FROM approval_history ah
        LEFT JOIN approval_flow af ON ah.flow_id = af.id
        WHERE ah.script_id = ? ORDER BY ah.version_number DESC, ah.created_at ASC
    ''', (script_id,))

def create_feedback_task(flow_id, script_id, version_number, modify_paragraph, feedback_content, assigned_to='策划师'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        task_id = execute_non_query('''
            INSERT INTO approval_task (flow_id, script_id, version_number, modify_paragraph, 
                                      feedback_content, status, assigned_to, created_at)
            VALUES (?, ?, ?, ?, ?, '待处理', ?, ?)
        ''', (flow_id, script_id, version_number, modify_paragraph, feedback_content, assigned_to, now))
        return task_id, None
    except Exception as e:
        return None, str(e)

def get_feedback_tasks_by_flow_id(flow_id):
    return execute_query('''
        SELECT id, flow_id, script_id, version_number, modify_paragraph, feedback_content, 
               status, assigned_to, created_at, completed_at
        FROM approval_task WHERE flow_id = ? ORDER BY created_at DESC
    ''', (flow_id,))

def get_feedback_tasks_by_script_id(script_id):
    return execute_query('''
        SELECT id, flow_id, script_id, version_number, modify_paragraph, feedback_content, 
               status, assigned_to, created_at, completed_at
        FROM approval_task WHERE script_id = ? ORDER BY created_at DESC
    ''', (script_id,))

def update_feedback_task_status(task_id, status, completed_at=None, operator_name='系统'):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    completed_at = completed_at or now
    try:
        execute_non_query('''
            UPDATE approval_task SET status=?, completed_at=? WHERE id=?
        ''', (status, completed_at, task_id))
        
        task = execute_query('''
            SELECT flow_id, script_id, version_number, modify_paragraph FROM approval_task WHERE id=?
        ''', (task_id,), fetch_one=True)
        
        if task and status == '已完成':
            add_approval_history(task[0], task[1], task[2], '反馈任务完成', '策划师', operator_name, f"完成段落: {task[3]}")
        
        return True, None
    except Exception as e:
        return False, str(e)

def approve_by_role(flow_id, script_id, version_number, action, operator_role, operator_name, remark=''):
    now = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    try:
        flow = get_approval_flow_by_id(flow_id)
        if not flow:
            return False, '审批流程不存在'
        
        if action == '确认':
            if operator_role == '主持人':
                update_approval_flow(flow_id, current_approver='策划师')
                add_approval_history(flow_id, script_id, version_number, '主持人确认', '主持人', operator_name, remark)
            elif operator_role == '策划师':
                update_approval_flow(flow_id, current_approver='客户')
                add_approval_history(flow_id, script_id, version_number, '策划师确认', '策划师', operator_name, remark)
        
        elif action == '催办':
            add_approval_history(flow_id, script_id, version_number, '催办', operator_role, operator_name, remark)
        
        elif action == '退回':
            update_approval_flow(flow_id, status='已退回', current_approver=operator_role)
            add_approval_history(flow_id, script_id, version_number, f'{operator_role}退回', operator_role, operator_name, remark)
            
            create_feedback_task(flow_id, script_id, version_number, '全文', remark)
        
        return True, None
    except Exception as e:
        return False, str(e)

def get_all_approval_flows():
    return execute_query('''
        SELECT af.id, af.script_id, af.version_number, af.status, af.current_approver, 
               hs.record_no, hs.bride_name, hs.groom_name, hs.host_name,
               af.created_at, af.customer_confirmed_at
        FROM approval_flow af
        LEFT JOIN host_script hs ON af.script_id = hs.id
        ORDER BY af.created_at DESC
    ''')

def export_approval_history(script_id=None, filename=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "审批历史"
    
    headers1 = ['审批ID', '脚本ID', '记录编号', '版本号', '操作', '操作角色', '操作人', '备注', '操作时间']
    for i, header in enumerate(headers1):
        cell = ws1.cell(row=1, column=i + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font.color = openpyxl.styles.colors.WHITE
        cell.alignment = Alignment(horizontal='center')
    
    if script_id:
        histories = get_approval_history_by_script_id(script_id)
    else:
        histories = execute_query('''
            SELECT ah.id, ah.script_id, hs.record_no, ah.version_number, ah.action, 
                   ah.operator_role, ah.operator_name, ah.remark, ah.created_at
            FROM approval_history ah
            LEFT JOIN host_script hs ON ah.script_id = hs.id
            ORDER BY ah.created_at DESC
        ''')
    
    row_idx = 2
    for h in histories:
        ws1.cell(row=row_idx, column=1, value=h[0])
        ws1.cell(row=row_idx, column=2, value=h[1])
        ws1.cell(row=row_idx, column=3, value=h[2] if len(h) > 2 else '')
        ws1.cell(row=row_idx, column=4, value=h[3] if len(h) > 3 else '')
        ws1.cell(row=row_idx, column=5, value=h[4] if len(h) > 4 else '')
        ws1.cell(row=row_idx, column=6, value=h[5] if len(h) > 5 else '')
        ws1.cell(row=row_idx, column=7, value=h[6] if len(h) > 6 else '')
        ws1.cell(row=row_idx, column=8, value=h[7] if len(h) > 7 else '')
        ws1.cell(row=row_idx, column=9, value=h[8] if len(h) > 8 else '')
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws1.column_dimensions[col].width = 20
    
    ws2 = wb.create_sheet(title="反馈任务")
    headers2 = ['任务ID', '审批流程ID', '脚本ID', '记录编号', '版本号', '修改段落', '反馈内容', '状态', '指派给', '创建时间', '完成时间']
    for i, header in enumerate(headers2):
        cell = ws2.cell(row=1, column=i + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        cell.font.color = openpyxl.styles.colors.WHITE
        cell.alignment = Alignment(horizontal='center')
    
    if script_id:
        tasks = get_feedback_tasks_by_script_id(script_id)
        record_no = execute_query('SELECT record_no FROM host_script WHERE id=?', (script_id,), fetch_one=True)
        record_no = record_no[0] if record_no else ''
    else:
        tasks = execute_query('''
            SELECT at.id, at.flow_id, at.script_id, hs.record_no, at.version_number, 
                   at.modify_paragraph, at.feedback_content, at.status, at.assigned_to, 
                   at.created_at, at.completed_at
            FROM approval_task at
            LEFT JOIN host_script hs ON at.script_id = hs.id
            ORDER BY at.created_at DESC
        ''')
        record_no = None
    
    row_idx = 2
    for t in tasks:
        ws2.cell(row=row_idx, column=1, value=t[0])
        ws2.cell(row=row_idx, column=2, value=t[1])
        ws2.cell(row=row_idx, column=3, value=t[2])
        if record_no:
            ws2.cell(row=row_idx, column=4, value=record_no)
            ws2.cell(row=row_idx, column=5, value=t[3])
            ws2.cell(row=row_idx, column=6, value=t[4])
            ws2.cell(row=row_idx, column=7, value=t[5])
            ws2.cell(row=row_idx, column=8, value=t[6])
            ws2.cell(row=row_idx, column=9, value=t[7])
            ws2.cell(row=row_idx, column=10, value=t[8])
            ws2.cell(row=row_idx, column=11, value=t[9])
        else:
            ws2.cell(row=row_idx, column=4, value=t[3] if len(t) > 3 else '')
            ws2.cell(row=row_idx, column=5, value=t[4] if len(t) > 4 else '')
            ws2.cell(row=row_idx, column=6, value=t[5] if len(t) > 5 else '')
            ws2.cell(row=row_idx, column=7, value=t[6] if len(t) > 6 else '')
            ws2.cell(row=row_idx, column=8, value=t[7] if len(t) > 7 else '')
            ws2.cell(row=row_idx, column=9, value=t[8] if len(t) > 8 else '')
            ws2.cell(row=row_idx, column=10, value=t[9] if len(t) > 9 else '')
            ws2.cell(row=row_idx, column=11, value=t[10] if len(t) > 10 else '')
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws2.column_dimensions[col].width = 20
    
    ws3 = wb.create_sheet(title="审批流程概览")
    headers3 = ['流程ID', '脚本ID', '记录编号', '版本号', '当前状态', '当前审批人', '客户查看时间', '确认结果', '确认人', '确认时间', '提交时间']
    for i, header in enumerate(headers3):
        cell = ws3.cell(row=1, column=i + 1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
        cell.font.color = openpyxl.styles.colors.WHITE
        cell.alignment = Alignment(horizontal='center')
    
    if script_id:
        flows = get_approval_flow_by_script_id(script_id)
    else:
        flows = execute_query('''
            SELECT af.id, af.script_id, hs.record_no, af.version_number, af.status, af.current_approver,
                   af.customer_view_time, af.customer_confirm_result, af.customer_confirmed_by,
                   af.customer_confirmed_at, af.created_at
            FROM approval_flow af
            LEFT JOIN host_script hs ON af.script_id = hs.id
            ORDER BY af.created_at DESC
        ''')
    
    row_idx = 2
    for f in flows:
        ws3.cell(row=row_idx, column=1, value=f[0])
        ws3.cell(row=row_idx, column=2, value=f[1])
        ws3.cell(row=row_idx, column=3, value=f[2] if len(f) > 2 else '')
        ws3.cell(row=row_idx, column=4, value=f[3] if len(f) > 3 else '')
        ws3.cell(row=row_idx, column=5, value=f[4] if len(f) > 4 else '')
        ws3.cell(row=row_idx, column=6, value=f[5] if len(f) > 5 else '')
        ws3.cell(row=row_idx, column=7, value=f[6] if len(f) > 6 else '')
        ws3.cell(row=row_idx, column=8, value=f[7] if len(f) > 7 else '')
        ws3.cell(row=row_idx, column=9, value=f[8] if len(f) > 8 else '')
        ws3.cell(row=row_idx, column=10, value=f[9] if len(f) > 9 else '')
        ws3.cell(row=row_idx, column=11, value=f[10] if len(f) > 10 else '')
        row_idx += 1
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws3.column_dimensions[col].width = 18
    
    if not filename:
        filename = f'审批历史_{datetime.datetime.now().strftime("%Y%m%d")}.xlsx'
    
    wb.save(filename)
    return filename, None